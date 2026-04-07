"""Generador de Excel profesional para la Gobernacion de Santander."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from backend.database.models import Case


# Colores institucionales
DARK_GREEN = "1A5276"
LIGHT_BLUE = "D6EAF8"
LIGHT_GRAY = "F2F3F4"
WHITE = "FFFFFF"
YELLOW = "FFF9C4"
GREEN = "C8E6C9"
RED = "FFCDD2"
ORANGE = "FFE0B2"
BLUE = "BBDEFB"
PINK = "F8BBD0"
PURPLE = "E1BEE7"

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=10)
HEADER_FILL = PatternFill(start_color=DARK_GREEN, end_color=DARK_GREEN, fill_type="solid")
DATA_FONT = Font(name="Calibri", size=9)
BORDER = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)

COLUMN_CONFIG = [
    ("_TIPO_ACTUACION", "TIPO", 14),
    ("RADICADO_23_DIGITOS", "RADICADO 23 DIGITOS", 28),
    ("RADICADO_FOREST", "RADICADO FOREST", 14),
    ("ABOGADO_RESPONSABLE", "ABOGADO", 25),
    ("ACCIONANTE", "ACCIONANTE", 30),
    ("ACCIONADOS", "ACCIONADOS", 35),
    ("VINCULADOS", "VINCULADOS", 30),
    ("DERECHO_VULNERADO", "DERECHO VULNERADO", 22),
    ("JUZGADO", "JUZGADO", 35),
    ("CIUDAD", "CIUDAD", 15),
    ("FECHA_INGRESO", "FECHA INGRESO", 13),
    ("ASUNTO", "ASUNTO", 40),
    ("PRETENSIONES", "PRETENSIONES", 45),
    ("OFICINA_RESPONSABLE", "OFICINA", 25),
    ("ESTADO", "ESTADO", 10),
    ("FECHA_RESPUESTA", "FECHA RESP.", 13),
    ("SENTIDO_FALLO_1ST", "FALLO 1ra", 18),
    ("FECHA_FALLO_1ST", "FECHA FALLO 1ra", 14),
    ("IMPUGNACION", "IMPUGN.", 9),
    ("QUIEN_IMPUGNO", "QUIEN IMPUGNO", 20),
    ("FOREST_IMPUGNACION", "FOREST IMPUGN.", 14),
    ("JUZGADO_2ND", "JUZGADO 2da", 30),
    ("SENTIDO_FALLO_2ND", "FALLO 2da", 14),
    ("FECHA_FALLO_2ND", "FECHA FALLO 2da", 14),
    ("INCIDENTE", "INCIDENTE", 10),
    ("FECHA_APERTURA_INCIDENTE", "FECHA INC.", 13),
    ("RESPONSABLE_DESACATO", "RESP. DESACATO", 20),
    ("DECISION_INCIDENTE", "DECISION INC.", 20),
    ("INCIDENTE_2", "INC.2", 8),
    ("FECHA_APERTURA_INCIDENTE_2", "FECHA INC.2", 13),
    ("RESPONSABLE_DESACATO_2", "RESP. DESAC.2", 20),
    ("DECISION_INCIDENTE_2", "DECISION INC.2", 20),
    ("INCIDENTE_3", "INC.3", 8),
    ("FECHA_APERTURA_INCIDENTE_3", "FECHA INC.3", 13),
    ("RESPONSABLE_DESACATO_3", "RESP. DESAC.3", 20),
    ("DECISION_INCIDENTE_3", "DECISION INC.3", 20),
    ("OBSERVACIONES", "OBSERVACIONES", 55),
]


def generate_excel(cases: list, output_path: str):
    """Generar archivo Excel profesional con 3 hojas."""
    wb = Workbook()

    _create_cover_sheet(wb, cases)
    _create_data_sheet(wb, cases)
    _create_stats_sheet(wb, cases)

    # Eliminar hoja por defecto si existe
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)


def _create_cover_sheet(wb: Workbook, cases: list):
    ws = wb.create_sheet("PORTADA", 0)

    # Titulo
    ws.merge_cells("B2:H2")
    cell = ws["B2"]
    cell.value = "GOBERNACION DE SANTANDER"
    cell.font = Font(name="Calibri", bold=True, size=22, color=DARK_GREEN)
    cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("B3:H3")
    cell = ws["B3"]
    cell.value = "SECRETARIA JURIDICA - GRUPO DE APOYO JURIDICO"
    cell.font = Font(name="Calibri", bold=True, size=14, color="555555")
    cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("B4:H4")
    cell = ws["B4"]
    cell.value = "CONSOLIDADO DE ACCIONES DE TUTELA 2026"
    cell.font = Font(name="Calibri", bold=True, size=16, color=DARK_GREEN)
    cell.alignment = Alignment(horizontal="center")

    # Metricas
    total = len(cases)
    activos = sum(1 for c in cases if (c.estado or "").upper() == "ACTIVO")
    inactivos = sum(1 for c in cases if (c.estado or "").upper() == "INACTIVO")
    concede = sum(1 for c in cases if "CONCEDE" in (c.sentido_fallo_1st or "").upper())
    niega = sum(1 for c in cases if "NIEGA" in (c.sentido_fallo_1st or "").upper())
    impugn = sum(1 for c in cases if (c.impugnacion or "").upper() == "SI")
    incid = sum(1 for c in cases if (c.incidente or "").upper() == "SI")

    tutelas = sum(1 for c in cases if (getattr(c, 'tipo_actuacion', '') or 'TUTELA') == 'TUTELA')
    incidentes_tipo = sum(1 for c in cases if (getattr(c, 'tipo_actuacion', '') or '') == 'INCIDENTE')

    metrics = [
        ("Total Carpetas", total, LIGHT_BLUE),
        ("Tutelas Unicas", tutelas, LIGHT_BLUE),
        ("Incidentes / Actuaciones Derivadas", incidentes_tipo, PURPLE),
        ("Casos Activos", activos, YELLOW),
        ("Casos Inactivos", inactivos, GREEN),
        ("Fallo: Concede (Desfavorable)", concede, RED),
        ("Fallo: Niega (Favorable)", niega, GREEN),
        ("Con Impugnacion", impugn, ORANGE),
        ("Con Incidente Desacato", incid, PURPLE),
    ]

    row = 7
    for label, value, color in metrics:
        ws.merge_cells(f"C{row}:E{row}")
        cell = ws[f"C{row}"]
        cell.value = label
        cell.font = Font(name="Calibri", bold=True, size=12)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.border = BORDER

        ws[f"F{row}"].value = value
        ws[f"F{row}"].font = Font(name="Calibri", bold=True, size=14)
        ws[f"F{row}"].alignment = Alignment(horizontal="center")
        ws[f"F{row}"].border = BORDER
        row += 1

    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 12


def _create_data_sheet(wb: Workbook, cases: list):
    ws = wb.create_sheet("TUTELAS")

    # Headers
    for col_idx, (_, header, width) in enumerate(COLUMN_CONFIG, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"

    # Datos
    for row_idx, case in enumerate(cases, 2):
        bg_color = WHITE if row_idx % 2 == 0 else LIGHT_GRAY

        for col_idx, (csv_col, _, _) in enumerate(COLUMN_CONFIG, 1):
            if csv_col == "_TIPO_ACTUACION":
                value = getattr(case, "tipo_actuacion", "TUTELA") or "TUTELA"
            else:
                attr = Case.CSV_FIELD_MAP.get(csv_col, "")
                value = getattr(case, attr, "") or ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

            # Colores semanticos
            if csv_col == "_TIPO_ACTUACION":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Calibri", bold=True, size=9)
                upper = value.strip().upper()
                if upper == "INCIDENTE":
                    cell.fill = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid")
                elif upper == "IMPUGNACION":
                    cell.fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

            elif csv_col == "ESTADO":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if "ACTIVO" == value.strip().upper():
                    cell.fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
                elif "INACTIVO" == value.strip().upper():
                    cell.fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")

            elif csv_col == "SENTIDO_FALLO_1ST":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                upper = value.strip().upper()
                if "CONCEDE" in upper:
                    cell.fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
                elif "NIEGA" in upper:
                    cell.fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
                elif "IMPROCEDENTE" in upper:
                    cell.fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")

            elif csv_col == "SENTIDO_FALLO_2ND":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                upper = value.strip().upper()
                if "CONFIRMA" in upper:
                    cell.fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
                elif "REVOCA" in upper:
                    cell.fill = PatternFill(start_color=PINK, end_color=PINK, fill_type="solid")

            elif csv_col in ("IMPUGNACION", "INCIDENTE", "INCIDENTE_2", "INCIDENTE_3"):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if value.strip().upper() == "SI":
                    cell.fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, size=9)

            elif csv_col == "RADICADO_23_DIGITOS":
                cell.font = Font(name="Courier New", size=8)

        ws.row_dimensions[row_idx].height = 40


def _create_stats_sheet(wb: Workbook, cases: list):
    ws = wb.create_sheet("ESTADISTICAS")

    # Titulo
    ws.merge_cells("B2:F2")
    cell = ws["B2"]
    cell.value = "ESTADISTICAS CONSOLIDADAS"
    cell.font = Font(name="Calibri", bold=True, size=16, color=DARK_GREEN)
    cell.alignment = Alignment(horizontal="center")

    total = len(cases)
    activos = sum(1 for c in cases if (c.estado or "").upper() == "ACTIVO")
    inactivos = sum(1 for c in cases if (c.estado or "").upper() == "INACTIVO")

    # Estado de casos
    row = 4
    ws.cell(row=row, column=2, value="ESTADO DE CASOS").font = Font(bold=True, size=11)
    for label, val in [("Total", total), ("ACTIVOS", activos), ("INACTIVOS", inactivos)]:
        row += 1
        ws.cell(row=row, column=2, value=label).font = DATA_FONT
        ws.cell(row=row, column=3, value=val).font = Font(bold=True, size=10)

    # Fallos
    row += 2
    ws.cell(row=row, column=2, value="FALLOS 1ra INSTANCIA").font = Font(bold=True, size=11)
    fallo_start = row + 1
    fallos = {"CONCEDE": 0, "NIEGA": 0, "IMPROCEDENTE": 0, "SIN FALLO": 0}
    for c in cases:
        f = (c.sentido_fallo_1st or "").upper()
        if "CONCEDE" in f:
            fallos["CONCEDE"] += 1
        elif "NIEGA" in f:
            fallos["NIEGA"] += 1
        elif "IMPROCEDENTE" in f:
            fallos["IMPROCEDENTE"] += 1
        else:
            fallos["SIN FALLO"] += 1

    for label, val in fallos.items():
        row += 1
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=3, value=val)

    # Grafico de fallos
    chart = BarChart()
    chart.title = "Distribucion de Fallos 1ra Instancia"
    chart.type = "col"
    chart.width = 18
    chart.height = 12
    data_ref = Reference(ws, min_col=3, min_row=fallo_start, max_row=row)
    cats_ref = Reference(ws, min_col=2, min_row=fallo_start, max_row=row)
    chart.add_data(data_ref)
    chart.set_categories(cats_ref)
    chart.legend = None
    ws.add_chart(chart, "E4")

    # Top ciudades
    row += 2
    ws.cell(row=row, column=2, value="TOP 10 CIUDADES").font = Font(bold=True, size=11)
    cities = {}
    for c in cases:
        city = (c.ciudad or "").strip()
        if city:
            cities[city] = cities.get(city, 0) + 1
    for city, count in sorted(cities.items(), key=lambda x: -x[1])[:10]:
        row += 1
        ws.cell(row=row, column=2, value=city)
        ws.cell(row=row, column=3, value=count)

    # Top abogados
    row += 2
    ws.cell(row=row, column=2, value="TOP 10 ABOGADOS").font = Font(bold=True, size=11)
    lawyers = {}
    for c in cases:
        lawyer = (c.abogado_responsable or "").strip()
        if lawyer:
            lawyers[lawyer] = lawyers.get(lawyer, 0) + 1
    for lawyer, count in sorted(lawyers.items(), key=lambda x: -x[1])[:10]:
        row += 1
        ws.cell(row=row, column=2, value=lawyer)
        ws.cell(row=row, column=3, value=count)

    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 10
