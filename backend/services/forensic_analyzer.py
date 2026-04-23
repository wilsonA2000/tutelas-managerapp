"""v5.2 — Forensic Document Analyzer (ingeniería inversa de cognición).

Emula el proceso mental de Claude/humano para clasificar documentos sin IA.
7 etapas: extracción → clasificación por contenido → entidades → identificadores
→ correlación → match DB → decisión.

Uso:
    from backend.services.forensic_analyzer import analyze_document, analyze_folder
    result = analyze_folder('/path/to/folder')
"""

from __future__ import annotations

import re
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

logger = logging.getLogger("tutelas.forensic")


# ═══════════════════════════════════════════════════════════════════════════
# ETAPA 2: Clasificación por contenido (no por nombre de archivo)
# ═══════════════════════════════════════════════════════════════════════════

DOC_TYPE_PATTERNS = {
    "EMAIL_OUTLOOK_TUTELA_ONLINE": [
        r"^Outlook\b", r"Tutela\s+en\s+Línea\s+(?:con\s+número|No\.?)\s*\d{7,}",
    ],
    "EMAIL_OUTLOOK_FORWARDED": [
        r"^Outlook\b", r"(?:Desde|De)[\s:]+.*@", r"Fecha\s+\w+",
    ],
    "EMAIL_MD_GMAIL": [
        r"^#\s+RV:", r"\*\*De:\*\*", r"\*\*Fecha:\*\*",
    ],
    # v5.2: documentos .docx generados por la Gobernación (respuestas, incidentes)
    "DOCX_RESPUESTA_FOREST": [
        r"(?:HONORABLE|Honorable|Señor|Señora)\s*\w*\s*JUEZ",
        r"(?:Proyect[óo]|Elabor[óo]|Revis[óo])\s*:",
        r"(?:FOREST|foret|Fores)\b",
    ],
    "DOCX_IMPUGNACION": [
        r"(?:IMPUGNACI[ÓO]N|IMPUGN[OA])", r"(?:Proyect[óo]|Elabor[óo])",
    ],
    "DOCX_DESACATO": [
        r"(?:INCIDENTE\s+DE\s+)?DESACATO", r"(?:Proyect[óo]|Elabor[óo])",
    ],
    "ESCRITO_TUTELA": [
        r"(?:Señor|Señora|Respetad[oa])\s*\w*\s*JUEZ",
        r"ACCIONANTE\s*:", r"ACCIONAD[OA]\s*:",
    ],
    "ACTA_REPARTO": [
        r"ACTA\s+DE\s+REPARTO", r"REPARTIDO\s+AL\s+JUZGADO",
    ],
    "AUTO_ADMISORIO": [
        r"(?:AVOCA|AVOC[OÓ])\s+(?:EL\s+)?CONOCIMIENTO",
        r"ADMIT[EO]\s+(?:LA\s+)?(?:PRESENTE\s+)?ACCI[ÓO]N",
    ],
    "SENTENCIA": [
        r"(?:FALLO|SENTENCIA|RESUELVE)\s*:", r"(?:CONCEDE|NIEGA|IMPROCEDENTE|AMPARA)",
    ],
    "AUTO_REQUERIMIENTO": [
        r"AUTO\s+(?:DE\s+)?REQUERIMIENTO", r"INCIDENTE\s+DE\s+DESACATO",
    ],
    "DERECHO_PETICION": [
        r"Derecho\s+de\s+Petici[óo]n", r"DERECHO\s+DE\s+PETICIÓN",
    ],
}


# v5.2: patrones ESPECÍFICOS para metadatos de emails .md de Gmail
MD_METADATA_PATTERNS = {
    "md_subject": r"^#\s+(.+?)$",  # primera línea H1
    "md_de": r"\*\*De:\*\*\s*(.+?)(?:\n|$)",
    "md_fecha": r"\*\*Fecha:\*\*\s*(.+?)(?:\n|$)",
    "md_caso": r"\*\*Caso:\*\*\s*(.+?)(?:\n|$)",
    "md_rad_from_subject": r"^#\s+.*?\b(20\d{2})[-\s]?0*(\d{2,5})\b",
    "md_rad_from_caso": r"\*\*Caso:\*\*\s*(20\d{2})[-\s]?0*(\d{2,5})",
}


def extract_md_metadata(text: str) -> dict:
    """v5.2: extrae metadatos de emails .md guardados por save_email_md()."""
    head = text[:1000]
    result = {}
    for key, pattern in MD_METADATA_PATTERNS.items():
        m = re.search(pattern, head, re.MULTILINE)
        if m:
            if key in ("md_rad_from_subject", "md_rad_from_caso"):
                result["rad_corto_md"] = f"{m.group(1)}-{int(m.group(2)):05d}"
            else:
                result[key] = m.group(1).strip()
    return result


# v5.2: patrones específicos para .docx de respuesta Gobernación
DOCX_RESPONSE_PATTERNS = {
    "abogado_proyecto": r"(?:Proyect[óo]|Elabor[óo])\s*:\s*([A-ZÁÉÍÓÚÑ][A-Za-záéíóúñ\s]{5,50}?)(?:\s+(?:[Rr]evis[óo]|[Aa]prob[óo]|[Aa]bogad|[Cc]ontratista|$))",
    "abogado_revisa": r"Revis[óo]\s*:\s*([A-ZÁÉÍÓÚÑ][A-Za-záéíóúñ\s]{5,50}?)(?:\s+(?:[Aa]prob[óo]|[Aa]bogad|$))",
    "abogado_aprueba": r"Aprob[óo]\s*:\s*([A-ZÁÉÍÓÚÑ][A-Za-záéíóúñ\s]{5,50}?)(?:\s+[Aa]bogad|$)",
    "forest_respuesta": r"(?:FOREST|Forest|forest)\s*(?:No\.?\s*)?:?\s*(\d{7,12})",
    "honorable_juez": r"Honorable\s+(?:Juez|Jueza)\s+(?:Doctor|Doctora|Dr\.?|Dra\.?)?\s*([A-ZÁÉÍÓÚÑ][A-Za-záéíóúñ\s]{5,50})?",
}


def extract_docx_response_metadata(text: str) -> dict:
    """v5.2: extrae metadatos específicos de .docx de respuesta (footer con abogado, FOREST)."""
    result = {}
    for key, pattern in DOCX_RESPONSE_PATTERNS.items():
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            if m.lastindex:
                result[key] = re.sub(r"\s+", " ", m.group(1).strip())[:80]
    return result


def classify_by_content(text: str, head_chars: int = 2000) -> list[tuple[str, int]]:
    """F1 Etapa 2: clasifica por estructura léxica.

    Returns: lista de (tipo, score) ordenada por score desc.
    Un documento puede tener varios tipos (ej. email que contiene sentencia).
    """
    head = (text or "")[:head_chars]
    if not head.strip():
        return [("SCAN_SIN_TEXTO", 100)]

    matches = []
    for doc_type, patterns in DOC_TYPE_PATTERNS.items():
        score = sum(1 for p in patterns if re.search(p, head, re.IGNORECASE | re.MULTILINE))
        if score > 0:
            matches.append((doc_type, score * 10))

    if not matches:
        return [("OTRO", 0)]
    return sorted(matches, key=lambda x: -x[1])


# ═══════════════════════════════════════════════════════════════════════════
# ETAPA 4: Extracción de TODOS los identificadores numéricos
# ═══════════════════════════════════════════════════════════════════════════

IDENTIFIER_PATTERNS = {
    "radicado_23d": r"\b(68\d{21})\b",
    "cc_accionante": r"(?:C\.?C\.?|[Cc]édula\s+(?:de\s+[Cc]iudadan[íi]a)?|Identificad[oa]\s+con\s+(?:documento|c[eé]dula)[:\s]+(?:No\.?\s*)?)\s*(\d{6,10})",
    "cc_generic": r"\b(\d{7,10})\b",  # fallback (validar con dígitos 6-10)
    "nuip_menor": r"(?:RC|Registro\s+Civil)\s+(?:No\.?\s*)?(\d{10,11})",
    "tutela_online": r"Tutela\s+(?:en\s+Línea\s+)?(?:con\s+número\s+|No\.?\s*)(\d{7,8})",
    "forest": r"(?:con\s+)?(?:[Nn][úu]mero\s+de\s+)?radicado\s+(20\d{9,11})",
    "expediente_disciplinario": r"Expediente\s+(?:No\.?\s*)?(\d{3,4}[-–]\d{2})",
    "nit": r"NIT[.\s:]+(\d{8,10}[-]?\d?)",
    "oficio_no": r"Oficio\s+(?:No\.?\s*)?(\d{1,5})",
    "acta_reparto_no": r"ACTA\s+DE\s+REPARTO(?:\s+\w+)?\s+No\.?\s*(\d+)",
}


def extract_all_identifiers(text: str) -> dict[str, list[str]]:
    """F1 Etapa 4: extrae todos los identificadores numéricos del texto."""
    result = {}
    text_clean = text or ""
    for name, pattern in IDENTIFIER_PATTERNS.items():
        matches = re.findall(pattern, text_clean, re.IGNORECASE)
        if matches:
            # Deduplicar preservando orden
            seen = set()
            unique = [m for m in matches if not (m in seen or seen.add(m))]
            result[name] = unique
    return result


# ═══════════════════════════════════════════════════════════════════════════
# ETAPA 3: Entidades específicas por tipo de documento
# ═══════════════════════════════════════════════════════════════════════════

ENTITY_PATTERNS = {
    "accionante_explicit": r"ACCIONANTE\s*:\s*([A-ZÁÉÍÓÚÑ][^\n]{5,80}?)(?:\s+(?:mayor|TIPO|identificad|C\.?C\.?|CÉDULA|\n))",
    "accionante_generic": r"ACCIONANTE\s*:\s*([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑa-záéíóúñ\s]{5,60})",
    "accionante_online": r"Accionante\s*:\s*([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑa-záéíóúñ\s]{5,60})\s+Identificad[oa]",
    "accionado": r"ACCIONAD[OA]S?\s*:\s*([A-ZÁÉÍÓÚÑ][^\n]{5,200})",
    "vinculado": r"VINCULAD[OA]S?\s*:\s*([A-ZÁÉÍÓÚÑ][^\n]{5,200})",
    "juzgado_reparto": r"REPARTIDO\s+AL\s+JUZGADO\s+_*(\d+|[A-Z]+)_*\s+(PROMISCUO|PENAL|CIVIL|LABORAL|PEQUEÑAS)[^\n]{5,80}",
    "juzgado_generic": r"Ju[zs]gado\s+(?:Promiscuo|Penal|Civil|Laboral|Primero|Segundo|Tercero|Cuarto|Quinto)\s+(?:Municipal|del\s+Circuito)[^\n]{5,80}",
    "ciudad_tutela_online": r"Ciudad\s*:\s*([A-ZÁÉÍÓÚÑ]+(?:\s+[A-ZÁÉÍÓÚÑ]+)*)",
    "correo_accionante": r"(?:[Cc]orreo\s+)?[Ee]lectr[óo]nico(?:\s+[Aa]ccionante)?\s*:?\s*([\w.\-]+@[\w.\-]+)",
    "derecho_vulnerado": r"(?:DERECHOS?\s+FUNDAMENTALES?|DERECHO\s+A\s+LA)\s+([A-ZÁÉÍÓÚÑ][^.\n]{5,200})",
    "fecha_hechos": r"(?:Desde|El\s+d[íi]a)\s+(?:el\s+)?(\d{1,2}\s+de\s+\w+\s+de\s+20\d{2})",
}


def extract_entities(text: str, doc_type: str = "") -> dict[str, str]:
    """F1 Etapa 3: extrae entidades según tipo de documento."""
    text_head = (text or "")[:5000]
    result = {}

    # Accionante (probar en orden de especificidad)
    for key in ("accionante_explicit", "accionante_online", "accionante_generic"):
        m = re.search(ENTITY_PATTERNS[key], text_head)
        if m:
            acc = re.sub(r"\s+", " ", m.group(1).strip())
            # Cortar en palabras-stop
            STOP = {"ACCIONADO", "ACCIONADA", "ACCIONANTE", "MAYOR", "IDENTIFICADO",
                    "IDENTIFICADA", "TIPO", "CÉDULA", "C.C", "CC"}
            tokens = acc.split()
            clean_tokens = []
            for t in tokens:
                if t.upper().strip(".,:;") in STOP:
                    break
                clean_tokens.append(t)
            if len(clean_tokens) >= 2:
                result["accionante"] = " ".join(clean_tokens).upper()[:60]
                break

    for key in ("accionado", "vinculado", "juzgado_reparto", "juzgado_generic",
                "ciudad_tutela_online", "correo_accionante", "derecho_vulnerado",
                "fecha_hechos"):
        m = re.search(ENTITY_PATTERNS[key], text_head, re.IGNORECASE)
        if m:
            value = re.sub(r"\s+", " ", m.group(1).strip())
            field_name = key.split("_")[0]  # accionado, vinculado, juzgado, ciudad, correo, derecho, fecha
            if field_name not in result:
                result[field_name] = value[:200]
    return result


# ═══════════════════════════════════════════════════════════════════════════
# Resultado del análisis
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class DocumentAnalysis:
    """Resultado forensic de un documento."""
    file_path: str
    filename: str
    doc_types: list[tuple[str, int]] = field(default_factory=list)
    identifiers: dict[str, list[str]] = field(default_factory=dict)
    entities: dict[str, str] = field(default_factory=dict)
    text_length: int = 0
    has_text: bool = False

    @property
    def primary_type(self) -> str:
        return self.doc_types[0][0] if self.doc_types else "OTRO"

    @property
    def accionante(self) -> Optional[str]:
        return self.entities.get("accionante")

    @property
    def cc_accionante(self) -> Optional[str]:
        ccs = self.identifiers.get("cc_accionante", [])
        return ccs[0] if ccs else None

    @property
    def rad23(self) -> Optional[str]:
        rads = self.identifiers.get("radicado_23d", [])
        return rads[0] if rads else None


def analyze_document(file_path: Path | str) -> DocumentAnalysis:
    """F1 Etapas 1-4: análisis completo de un documento individual."""
    path = Path(file_path)
    analysis = DocumentAnalysis(file_path=str(path), filename=path.name)
    ext = path.suffix.lower()

    # Etapa 1: extraer texto. v5.5: leer documentos ampliamente para cosechar
    # señales deterministas en footers (Proyectó:), sellos en anexos y watermarks.
    text = _extract_text(path, max_chars=150000)
    analysis.text_length = len(text)
    analysis.has_text = len(text.strip()) > 50

    if not analysis.has_text:
        analysis.doc_types = [("SCAN_SIN_TEXTO", 100)]
        # v5.2: aunque no haya texto, el FILENAME puede dar pistas (ej. "2026-00014 AutoCierre.pdf")
        fn_identifiers = extract_all_identifiers(path.name)
        if fn_identifiers:
            analysis.identifiers = fn_identifiers
        return analysis

    # Etapa 2: clasificar
    analysis.doc_types = classify_by_content(text)

    # Etapa 4: identificadores (texto + filename como complemento)
    analysis.identifiers = extract_all_identifiers(text + " " + path.name)

    # Etapa 3: entidades
    analysis.entities = extract_entities(text, doc_type=analysis.primary_type)

    # v5.2: extractors específicos por extensión
    if ext == ".md":
        md_meta = extract_md_metadata(text)
        if md_meta:
            analysis.entities.update({f"md_{k.replace('md_','')}": v for k, v in md_meta.items()})
            # Si el .md tiene rad_corto en subject, agregarlo a identifiers
            if "rad_corto_md" in md_meta:
                analysis.identifiers.setdefault("rad_corto_from_md", []).append(md_meta["rad_corto_md"])

    if ext in (".docx", ".doc"):
        docx_meta = extract_docx_response_metadata(text)
        if docx_meta:
            analysis.entities.update(docx_meta)

    return analysis


def analyze_text_blob(
    text: str,
    attachment_filenames: list[str] | None = None,
    forwarded_blocks: list[str] | None = None,
) -> DocumentAnalysis:
    """v5.4.4: análisis forense sobre texto plano + filenames (sin file IO, sin DB).

    Diseñado para el monitor Gmail: recibe subject+body del email + nombres de
    adjuntos detectados (sin descargar aún) + bloques forwarded opcionales para
    priorizar el bloque raíz.

    Args:
        text: subject + body completo del email.
        attachment_filenames: nombres de archivos adjuntos (se añaden al texto
            de identificadores porque PDFs de juzgado frecuentemente traen el
            rad23 en el nombre, ej. "8001333300820230026700_Auto.pdf").
        forwarded_blocks: si se proveen, se priorizan los 2 primeros para
            extracción de entidades (el bloque raíz tiene menor probabilidad
            de contaminación cruzada).

    Returns:
        DocumentAnalysis con identifiers + entities. text_length = len(text).
    """
    analysis = DocumentAnalysis(file_path="<email-blob>", filename="<email>")

    text_full = text or ""
    analysis.text_length = len(text_full)
    analysis.has_text = len(text_full.strip()) > 20

    if not analysis.has_text and not attachment_filenames:
        analysis.doc_types = [("EMPTY_BLOB", 100)]
        return analysis

    # Etapa 2: clasificar (solo sobre el cuerpo del email)
    if analysis.has_text:
        analysis.doc_types = classify_by_content(text_full)
    else:
        analysis.doc_types = [("EMAIL", 50)]

    # Etapa 4: identificadores.
    # Concatenamos text + filenames para que los regex de rad23/CC/FOREST
    # también minen los nombres de archivo adjunto.
    search_text = text_full
    if attachment_filenames:
        search_text = search_text + "\n" + "\n".join(attachment_filenames)
    analysis.identifiers = extract_all_identifiers(search_text)

    # Etapa 3: entidades. Priorizar bloque raíz si se provee.
    if forwarded_blocks:
        # Tomar los 2 primeros bloques (más cercanos al remitente original)
        priority_text = "\n".join(forwarded_blocks[:2])
        analysis.entities = extract_entities(priority_text)
        # Fallback: si no se encontró accionante en prioridad, buscar en todo
        if not analysis.entities.get("accionante"):
            fallback = extract_entities(text_full)
            for k, v in fallback.items():
                analysis.entities.setdefault(k, v)
    else:
        analysis.entities = extract_entities(text_full)

    return analysis


def _extract_text(path: Path, max_chars: int = 150000) -> str:
    """Extrae texto de PDF/DOCX/DOC/MD/TXT/XLSX + footers DOCX + OCR imágenes.

    v5.5: lee todas las páginas del PDF (antes limitaba a 3) para cosechar
    señales deterministas en anexos, firmas finales y sellos tardíos.
    """
    ext = path.suffix.lower()
    try:
        if ext == ".pdf":
            import fitz
            doc = fitz.open(str(path))
            text = ""
            for page_num in range(len(doc)):
                text += doc[page_num].get_text()
                if len(text) > max_chars:
                    break
            doc.close()
            return text[:max_chars]

        elif ext in (".md", ".txt"):
            return path.read_text(encoding="utf-8", errors="replace")[:max_chars]

        elif ext == ".docx":
            # v5.2: leer body + footers (footers tienen "Proyectó:" del abogado Gobernación)
            import zipfile
            from xml.etree import ElementTree as ET
            parts = []
            with zipfile.ZipFile(str(path)) as z:
                try:
                    with z.open("word/document.xml") as f:
                        tree = ET.parse(f)
                        parts.append(" ".join(t.text or "" for t in tree.iter() if t.text))
                except KeyError:
                    pass
                for footer in ("word/footer1.xml", "word/footer2.xml", "word/footer3.xml"):
                    try:
                        with z.open(footer) as f:
                            tree = ET.parse(f)
                            parts.append(" ".join(t.text or "" for t in tree.iter() if t.text))
                    except KeyError:
                        break
            return " ".join(parts)[:max_chars]

        elif ext == ".doc":
            # v5.2: legacy binario — cascada fitz → antiword → olefile scan
            for method in (_extract_doc_fitz, _extract_doc_antiword, _extract_doc_olefile):
                try:
                    r = method(path, max_chars)
                    if r and r.strip():
                        return r
                except Exception:
                    continue
            return ""

        elif ext == ".xlsx":
            # v5.2: reportes Excel
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(path), read_only=True, data_only=True)
                parts = []
                for sheet in wb.worksheets[:3]:
                    for row in sheet.iter_rows(max_row=30, values_only=True):
                        parts.append(" | ".join(str(c) for c in row if c is not None))
                wb.close()
                return "\n".join(parts)[:max_chars]
            except ImportError:
                return ""

        elif ext in (".png", ".jpg", ".jpeg"):
            # v5.2: imágenes con OCR si disponible
            try:
                from backend.extraction.ocr_extractor import extract_image_ocr, is_tesseract_available
                if is_tesseract_available():
                    return (extract_image_ocr(str(path)).text or "")[:max_chars]
            except Exception:
                pass
            return ""

    except Exception as e:
        logger.warning("Fallo extract_text %s: %s", path.name, e)
    return ""


def _extract_doc_fitz(path: Path, max_chars: int) -> str:
    import fitz
    doc = fitz.open(str(path))
    text = doc[0].get_text() if len(doc) else ""
    doc.close()
    return text[:max_chars]


def _extract_doc_antiword(path: Path, max_chars: int) -> str:
    import subprocess, shutil as _sh
    if not _sh.which("antiword"):
        return ""
    r = subprocess.run(["antiword", str(path)], capture_output=True, timeout=15)
    return r.stdout.decode("utf-8", errors="replace")[:max_chars]


def _extract_doc_olefile(path: Path, max_chars: int) -> str:
    try:
        import olefile
        import re as _re
        with olefile.OleFileIO(str(path)) as ole:
            for stream in ("WordDocument", "1Table"):
                if ole.exists(stream):
                    raw = ole.openstream(stream).read()
                    chunks = _re.findall(rb"[\x20-\x7e\xc0-\xff]{6,}", raw)
                    decoded = b" ".join(chunks[:300]).decode("utf-8", errors="replace")
                    if len(decoded) > 100:
                        return decoded[:max_chars]
    except ImportError:
        pass
    except Exception:
        pass
    return ""
